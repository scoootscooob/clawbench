"""Recursive workspace search verifier."""

from __future__ import annotations

import sys
from pathlib import Path

EXCLUDE_FRAGMENTS = (
    "verify_", "/.git/", "/.openclaw/",
    "BOOTSTRAP.md", "IDENTITY.md", "AGENTS.md",
    "USER.md", "SOUL.md", "HEARTBEAT.md",
)
TEXT_SUFFIXES = (".md", ".txt", ".json", ".yaml", ".yml", ".csv", ".log",
                  ".jsonl", ".html", ".sh", ".py")


def iter_workspace_text_files(root: Path = Path(".")):
    for path in root.rglob("*"):
        if not path.is_file():
            continue
        sp = str(path)
        if any(frag in sp for frag in EXCLUDE_FRAGMENTS):
            continue
        if path.suffix.lower() not in TEXT_SUFFIXES:
            continue
        try:
            yield path, path.read_text(encoding="utf-8", errors="ignore")
        except Exception:
            continue


def workspace_blob() -> str:
    return "\n".join(text for _, text in iter_workspace_text_files())


import json

def main() -> int:
    expected = json.loads(Path(".expected_totals.json").read_text())
    expected_strs = {r: str(t) for r, t in expected.items()}

    # First try the structured xlsx
    try:
        import openpyxl
        for path in Path(".").rglob("*.xlsx"):
            if "verify_" in str(path):
                continue
            try:
                wb = openpyxl.load_workbook(path, data_only=True)
            except Exception:
                continue
            flat = []
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                for row in ws.iter_rows(values_only=True):
                    for cell in row:
                        if cell is not None:
                            flat.append(str(cell))
            blob = " ".join(flat)
            if all(r in blob for r in expected.keys()) and all(t in blob for t in expected_strs.values()):
                print(f"PASS: rollup totals found in {path}")
                return 0
    except ImportError:
        pass

    # Fall back to any text file
    blob = workspace_blob()
    if all(r in blob for r in expected.keys()) and all(t in blob for t in expected_strs.values()):
        print("PASS: rollup totals found in workspace text")
        return 0
    print(f"FAIL: regional totals not found anywhere. Expected: {expected}")
    return 1


if __name__ == "__main__":
    sys.exit(main())
